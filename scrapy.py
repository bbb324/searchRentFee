import requests
import math
from bs4 import BeautifulSoup
import re # 正则
from bokeh.io import output_notebook, show # 引入初始化
from bokeh.plotting import figure #引入图表界面
from bokeh.models import ColumnDataSource, HoverTool, FactorRange, LabelSet
from bokeh.models.annotations import Label # 引入备注
from bokeh.sampledata.autompg import autompg_clean as df
from bokeh.sampledata.commits import data
from bokeh.transform import factor_cmap

from bokeh.models.glyphs import VBar

from datetime import datetime
import time
from openpyxl import Workbook
from openpyxl.cell.cell import WriteOnlyCell # 写excel
from openpyxl.utils.dataframe import dataframe_to_rows



# 爬取特定区域内挂牌7天以内的房源
def grab(url, last_day):
    r = requests.get(url)
    soup = BeautifulSoup(r.content, 'lxml')
    g_data = soup.find_all("div", {"class", "info clear"})
    
    result = []

    for item in g_data:
        day = item.contents[3].text.split('/ ')[2]
        
        if day == '刚刚发布' or int(re.match('\d+', day).group()) < last_day:
            try:
                yard = item.contents[1].find_all('a')[0].text
            except:
                pass
            try:
                price = item.contents[5].find_all('span')[1].text
            except:
                pass
            try:
                address = item.contents[4].find_all('span', {'class', 'subway'})[0].text
            except:
                pass
            try:
                time = item.contents[3].text.split('/ ')[2]
            except:
                pass
            result.append({
                'yard'  : yard,
                'price' : price,
                'time'  : time,
                'url'   : url
            })    
    return result;

            #print(yard+ '|' + price + '|' + time + '|') #打印出 小区，价格，地址
    

#先获取总页数，再以循环的方式获取列表
def getTotal(day):
    # page = 5 # 爬的总页数
    # count = 1
    # area_list = [
    #     'https://sz.lianjia.com/ershoufang/luohuqu/pg', # 罗湖
    #     'https://sz.lianjia.com/ershoufang/futianqu/pg', # 福田
    #     'https://sz.lianjia.com/ershoufang/nanshanqu/pg', # 南山
    #     'https://sz.lianjia.com/ershoufang/baoanqu/pg'] # 宝安
    # total = [];    
    # for link in area_list:
    #     while count<=page:
    #         href   = link + str(count) +"co32a3a4/"
    #         #info   = grab(href, day)
    #         if(len(info) != 0):
    #            total.append(info) 
    #         count += 1
    #     count = 1    
    # print(total)    
    page = 5 # 爬的总页数
    count = 1
    area_list = {
        'luohu'   : 'https://sz.lianjia.com/ershoufang/luohuqu/pg',
        'futian'  : 'https://sz.lianjia.com/ershoufang/futianqu/pg',
        'nanshan' : 'https://sz.lianjia.com/ershoufang/nanshanqu/pg',
        'baoan'   : 'https://sz.lianjia.com/ershoufang/baoanqu/pg'
    }
    total = [];    
    for link in area_list:
        while count<=page:
            href   = link + str(count) +"co32a3a4/"
            #info   = grab(href, day)
            if(len(info) != 0):
               total.append(info) 
            count += 1
        count = 1    
    print(total)    
# 爬取 7天内 新上架的房源
getTotal(7)




def grab_area_house_number():
    title = '各区房源数量'
    nums = []
    area_list = [
        'https://sz.lianjia.com/ershoufang/luohuqu/a3a4', # 罗湖
        'https://sz.lianjia.com/ershoufang/futianqu/a3a4', # 福田
        'https://sz.lianjia.com/ershoufang/nanshanqu/a3a4', # 南山
        'https://sz.lianjia.com/ershoufang/baoanqu/a3a4'] # 宝安
    
    for area in area_list:
        page = requests.get(area)
        soup = BeautifulSoup(page.content, 'lxml')
        total = soup.find('h2', {'class', 'total fl'})
        number = int(total.find('span').text)
        nums.append(number)
    
    label_name = ['罗湖','福田','南山','宝安']

    diagram = figure(x_range     = label_name, 
                    y_range      = [0, 2000], 
                    plot_height  = 250, 
                    title        = title, 
                    tools        = 'wheel_zoom',
                    y_axis_label = '房源数量')

    source = ColumnDataSource(data=dict(
        x     = label_name,
        top   = nums,
        names = nums,
        color = ['#67b7dc', '#b9783f','#84b761', '#cd82ad']))

    labels = LabelSet(source=source,
                        x           = 'x',
                        y           = 'top',
                        text        = 'names', 
                        level       = 'glyph', 
                        text_align  = 'center', 
                        render_mode = 'canvas')
    diagram.vbar(source=source, 
                    x      = 'x', 
                    top    = 'top', 
                    width  = 0.5, 
                    bottom =  0, 
                    color  = 'color')

    # 添加顶部具体数据
    diagram.add_layout(labels)
    

    show(diagram)
    
# 爬取特定区域的房源总数 （70~90，90~110 不考虑价格）
#grab_area_house_number()


# 30 天内的成交数据
def getDoneTransaction():
    url    = 'https://sz.lianjia.com/chengjiao/nanshanzhongxin/'
    r      = requests.get(url)
    soup   = BeautifulSoup(r.content, 'lxml')
    g_data = soup.find_all("ul", {"class", "listContent"})

    for item in g_data[0].find_all('li'):
        try:
            yard    = item.find_all('div', {'class', 'title'})[0].text
        except:
            pass
        try:
            href    = item.find_all('a')[0].attrs['href']
            result  = getTransactionDate(href)
        except:
            pass
        print(yard + '|' + result['total']+ '|' + result['date'])    



# 成交具体详情（小区，总价，日期）
def getTransactionDate(url):
    req         = requests.get(url)
    soup        = BeautifulSoup(req.content, 'lxml')
    detail      = soup.find_all("div", {"class", "wrapper"})[0]
    total_ele   = soup.find_all("span", {"class", "dealTotalPrice"})[0]


    date        = detail.find_all('span')[0].text.split(' ')[0] # 成交日期
    total_value = total_ele.find_all('i')[0].text + '万' # 成交金额

    return ({'total': total_value, 'date': date})


#getDoneTransaction()

def getUploadInfo(url):

    req         = requests.get(url)
    soup        = BeautifulSoup(req.content, 'lxml')
    priceInfo   = soup.find_all("div", {"class", "price"})[0]
    total       = priceInfo.find_all('span', {'class', 'total'})[0].text + '万'
    unit        = priceInfo.find_all('span', {'class', 'unitPriceValue'})[0].text
    
    # 房子信息
    houseInfo   = soup.find_all("div", {"class", "houseInfo"})[0]
    room        = houseInfo.find_all('div', {'class', 'mainInfo'})[0].text

    area        = soup.find_all('div', {'class', 'area'})[0]
    size        = area.find_all('div', {'class', 'mainInfo'})[0].text + '平米'
    age         = area.find_all('div', {'class', 'subInfo'})[0].text
    direction   = soup.find_all('div', {'class', 'type'})[0].find_all('div')[0].text

    geoInfo     = soup.find_all('div', {'class', 'aroundInfo'})[0]
    community   = geoInfo.find_all('a')[0].text

    # 挂牌时间
    date = soup.find_all('div', {'class', 'transaction'})[0].find_all('span')[1].text
    return ({
                'total'     : total,      # 总价
                'unit'      : unit,       # 单价
                'room'      : room,       # 房屋结构
                'direction' : direction,  # 房屋朝向
                'community' : community,  # 房屋所属小区
                'age'       : age,        # 房龄
                'size'      : size,       # 房屋大小
                'date'      : date        # 挂牌时间
            })





def getTotalTransactionByArea(priceRange, deadline):
    area_list    = ['bagualing', 'baihua','chegongmian','chiwei','futianbaoshuiqu']
    baseUrl = 'https://sz.lianjia.com/ershoufang/'

    # tmpUrl = 'https://sz.lianjia.com/ershoufang/baihua/co21'
    # r      = requests.get(tmpUrl)
    # soup   = BeautifulSoup(r.content, 'lxml')
    # g_data = soup.find_all('div', {'class', 'info clear'})


    info = []

    
    for location in area_list:
        href   = baseUrl + location + '/co21'
        r      = requests.get(href)
        soup   = BeautifulSoup(r.content, 'lxml')
        g_data = soup.find_all('div', {'class', 'info clear'})
        
        
        for item in g_data:
            price = float(item.contents[5].find_all('span')[0].text)
            result  = getUploadInfo(item.find_all('a')[0].attrs['href'])
            print(result)
            if price >= priceRange[0] and price <= priceRange[1] and deadline <= result['date']:

                result['link']       = item.find_all('a')[0].attrs['href']
                result['uploadTime'] = item.contents[3].text.split('/ ')[2]
                info.append(result)
        if len(info) > 0:
            write_excel(info, location)

def write_excel(info, location):
    wb = Workbook()
    ws = wb.active
    ws.append([
        'total',       # 总价
        'unit',        # 单价
        'room'         # 户型
        'direction'    # 朝向
        'community'    # 小区
        'age'          # 房龄
        'size'         # 平方米
        'link'         # 链接
        'uploadTime'   # 上传时间
    ])
    
    for key in range(len(info)):
        ws.cell(row = (key + 2), column = 1).value = info[key]['total']
        ws.cell(row = (key + 2), column = 2).value = info[key]['unit']
        ws.cell(row = (key + 2), column = 3).value = info[key]['room']
        ws.cell(row = (key + 2), column = 4).value = info[key]['direction']
        ws.cell(row = (key + 2), column = 5).value = info[key]['community']
        ws.cell(row = (key + 2), column = 6).value = info[key]['age']
        ws.cell(row = (key + 2), column = 7).value = info[key]['size']
        ws.cell(row = (key + 2), column = 8).value = info[key]['link']
        ws.cell(row = (key + 2), column = 9).value = info[key]['uploadTime']
    wb.save(location + '.xlsx')
    

# 获取福田区各位置总价范围600~800万，2018-04-01之后上的房源
#getTotalTransactionByArea([600,800], '2018-04-01')




def visualize():
    # 折线图画法
    p = figure(plot_width=800, plot_height=300, title="abc")
    p.multi_line(xs=[[1,2,3], [2,3,4]], ys=[[6,7,2], [4,5,7]], color=['red', 'black'])

    

    # 柱状图
    plot = figure(plot_width=800, plot_height=300, title='rect')
    plot.vbar(x=['a','b','c'], width=0.5, bottom=0, top=[1,2,3], color=['red', 'brown', 'navy'])

    # pandas, source 是坐标,可以直接引入外部数据通过 bokeh.sampledata.autompg
    source = ColumnDataSource(data={
        'x': [1,2,3,4],
        'y': [3,6,4,2],
        })
    p_panda = figure(plot_width=800, plot_height=300)
    p_panda.circle('x', 'y', source=source)
    
    # x轴为label的柱状图展示
    fruites=['a','b','c','d']
    p_fruits = figure(x_range=fruites, plot_height=250, title='fruit')
    p_fruits.vbar(x=fruites, top=[5,2,3,6], width=0.5)
    p_fruits.xgrid.grid_line_color= 'red'
    p_fruits.y_range.start = 0


    #点图加注释LabelSet
    source = ColumnDataSource(data=dict(
        temp=[166, 171, 172, 168, 174, 162],
        pressure=[165, 189, 220, 141, 260, 174],
        names=['A', 'B', 'C', 'D', 'E', 'F']))

    p_anotation = figure(x_range=(160, 175))
    p_anotation.scatter(x='temp', y='pressure', size=8, source=source)
    p_anotation.xaxis.axis_label = 'Temperature (C)'
    p_anotation.yaxis.axis_label = 'Pressure (lbs)'

    labels = LabelSet(x='temp', y='pressure', text='names', level='glyph',
                  x_offset=5, y_offset=5, source=source, render_mode='canvas')

    p_anotation.add_layout(labels)
    show(p_anotation)

#visualize() 

    # 按品种为单位，三年的数量柱状图
def groupChart():
    fruits = ['Apples', 'Pears', 'Nectarines', 'Plums', 'Grapes', 'Strawberries']
    years = ['2015', '2016', '2017']
    data = {'fruits' : fruits,
            '2015'   : [2, 1, 4, 3, 2, 4],
            '2016'   : [5, 3, 3, 2, 4, 6],
            '2017'   : [3, 2, 4, 4, 5, 3]}

    # this creates [ ("Apples", "2015"), ("Apples", "2016"), ("Apples", "2017"), ("Pears", "2015), ... ]
    x = [ (fruit, year) for fruit in fruits for year in years ]
    counts = sum(zip(data['2015'], data['2016'], data['2017']), ()) # like an hstack

    source = ColumnDataSource(data=dict(x=x, counts=counts))

    p = figure(x_range=FactorRange(*x), plot_height=250, title="Fruit Counts by Year")

    p.vbar(x='x', top='counts', width=0.9, source=source)

    p.y_range.start = 0
    p.x_range.range_padding = 0.1
    p.xaxis.major_label_orientation = 1
    p.xgrid.grid_line_color = None
    show(p)
#groupChart()

# 混合图形，包含柱状图，折线图，并且按“季度”为group展开，折线图为4季度平均值连线
def mixedCategory():
    factors = [("Q1", "jan"), ("Q1", "feb"), ("Q1", "mar"),
           ("Q2", "apr"), ("Q2", "may"), ("Q2", "jun"),
           ("Q3", "jul"), ("Q3", "aug"), ("Q3", "sep"),
           ("Q4", "oct"), ("Q4", "nov"), ("Q4", "dec")]
    p = figure(x_range=FactorRange(*factors), plot_height=250)
    x = [ 10, 12, 16, 9, 10, 8, 12, 13, 14, 14, 12, 16 ]
    p.vbar(x=factors, top=x, width=0.9, alpha=0.5)       
    qs, aves = ["Q1", "Q2", "Q3", "Q4"], [(10+12+16)/3, (9+10+8)/3, (12+13+14)/3, (14+12+16)/3]
    p.line(x=qs, y=aves, color="red", line_width=3)
    p.circle(x=qs, y=aves, line_color="red", fill_color="white", size=10)

    p.y_range.start = 0
    p.x_range.range_padding = 0.1
    p.xgrid.grid_line_color = None
    show(p)
#mixedCategory()









