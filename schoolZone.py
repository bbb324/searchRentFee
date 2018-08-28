import requests
import math
from bs4 import BeautifulSoup
import re # 正则
import time
from blacklist import blacklist
from openpyxl import Workbook

link_list = [
    'http://www.szxuexiao.com/PrimarySchool/LuoHuShengYiJiXiaoXue.html',    # 罗湖学校列表
    'http://www.szxuexiao.com/PrimarySchool/FuTianShengYiJiXiaoXue.html',   # 福田学校列表
    'http://www.szxuexiao.com/PrimarySchool/NaShanShengYiJiXiaoXue.html',   # 南山学校列表
]

dist_list = [
        'http://bsy.sz.bendibao.com/bsyDetail/606150.html', #罗湖住房类表
        'http://bsy.sz.bendibao.com/bsyDetail/611764.html', #福田住房列表
        'http://bsy.sz.bendibao.com/bsyDetail/616443.html', #南山住房列表
]
tmp = [{'total': '420万', 'unit': '52500元/平米', 'room': '2室2厅', 'direction': '南 北', 'community': '泰然广场', 'age': '1998年建/板楼', 'size': '80平米平米', 'date': '2016-10-30', 'url': 'https://sz.lianjia.com/ershoufang/105100484153.html', 'schoolName': '下沙小学'}, {'total': '530万', 'unit': '54640元/平米', 'room': '3室2厅', 'direction': '南 北', 'community': '泰然广场', 'age': '1998年建/板楼', 'size': '97平米平米', 'date': '2018-01-13', 'url': 'https://sz.lianjia.com/ershoufang/105101188890.html', 'schoolName': '下沙小学'}, {'total': '540万', 'unit': '55556元/平米', 'room': '3室2厅', 'direction': '南', 'community': '泰然广场', 'age': '1997年建/板楼', 'size': '97.2平米平米', 'date': '2017-12-23', 'url': 'https://sz.lianjia.com/ershoufang/105101159403.html', 'schoolName': '下沙小学'}]

wb = Workbook()
ws = wb.active

# 获得省一级小学列表名单
def getTopElementrySchool(area, range):
    if area == 'luohu':
        url = link_list[0]
    elif area == 'futian':
        url = link_list[1]
    elif area == 'nanshan':
        url = link_list[2]

    r = requests.get(url)
    soup = BeautifulSoup(r.content, 'lxml')
    group = soup.find_all("div", {"class", "doc_list"})
    ul = group[0].find_all('ul', {'class', 'detail_nursery'})
    img = ul[0].find_all('img');
    info = []
    for item in ul:
        text = item.find_all('img')[0].attrs['alt']
        schoolList = regexGetSchoolName(text, area) # 字符串截取获取小学名称
        info.append(schoolList)  
    getLocationName(info, area, range)


def getLocationName(info, area, range):
    if area == 'luohu':
        url = dist_list[0]
    elif area == 'futian':
        url = dist_list[1]
    elif area == 'nanshan':
        url = dist_list[2]
   
    r = requests.get(url)
    soup = BeautifulSoup(r.content, 'lxml')
    group = soup.find_all("div", {"class", "checkbox1"})
    
    for name in info:
        for dist in group:
            titles = dist.find_all('strong')
            for title in titles:
                if name in title.text or ('石厦' in name and '石厦' in title.text):
                    dicts = {
                        'area': area,
                        'name': name,
                        'community': getCommunityName(title.parent.nextSibling)
                    }
                    getApartmentInfo(dicts, range)

# 获取学校真实名称
def regexGetSchoolName(text, area):
    if area == 'futian':
            schoolName = text;
            if '深圳福田' in text:
                schoolName = text.split('深圳福田')[1]
    return schoolName

# 获取对应学校所包含的小区信息
def getCommunityName(ele):
    nameList = ele.find(id='s_list')
    
    if nameList != None:
        return nameList.text.split('，')
    elif ele != None and '学区划分' in ele.text:
        return ele.text.split('：')[1].split('，')
    # 荔园外语东校区
    elif ele.nextSibling != None and '学区划分' in ele.nextSibling.text:
        return ele.nextSibling.text.split('：')[1].split('，')
    # 荔园小学西校区
    elif ele.nextSibling.nextSibling.nextSibling.find(id='s_list') != None:
        return ele.nextSibling.nextSibling.nextSibling.find(id='s_list').text.split('，')
    else:
        print('no')

# 得到房源具体信息准备录入excel
def getUploadInfo(url, schoolName):
    req         = requests.get(url)
    soup        = BeautifulSoup(req.content, 'lxml')
    priceInfo   = soup.find_all("div", {"class", "price"})[0]
    total       = priceInfo.find_all('span', {'class', 'total'})[0].text + '万'
    unit        = priceInfo.find_all('span', {'class', 'unitPriceValue'})[0].text
    
    # 房子信息
    houseInfo   = soup.find_all("div", {"class", "houseInfo"})[0]
    room        = houseInfo.find_all('div', {'class', 'mainInfo'})[0].text

    area        = soup.find_all('div', {'class', 'area'})[0]
    size        = area.find_all('div', {'class', 'mainInfo'})[0].text + '平米'
    age         = area.find_all('div', {'class', 'subInfo'})[0].text
    direction   = soup.find_all('div', {'class', 'type'})[0].find_all('div')[0].text
    geoInfo     = soup.find_all('div', {'class', 'aroundInfo'})[0]
    community   = geoInfo.find_all('a')[0].text

    # 挂牌时间
    date = soup.find_all('div', {'class', 'transaction'})[0].find_all('span')[1].text
    return ({
                'total'     : total,      # 总价
                'unit'      : unit,       # 单价
                'room'      : room,       # 房屋结构
                'direction' : direction,  # 房屋朝向
                'community' : community,  # 房屋所属小区
                'age'       : age,        # 房龄
                'size'      : size,       # 房屋大小
                'date'      : date,       # 挂牌时间
                'url'       : url,        # 房源链接
                'schoolName': schoolName  # 学校名称
            })


# 获取每个小区的价格
def getApartmentInfo(param, range):
    communities = param['community']
    houseInfo = []
    for unit in communities:
        url = 'https://sz.lianjia.com/ershoufang/bp'+str(range[0]) + 'ep'+str(range[1])+'rs'+ unit
        r = requests.get(url)
        soup = BeautifulSoup(r.content, 'lxml')

        # 打印学校名称和对应小区名称
        print(param['name'], unit)
        group = soup.find("h2", {"class", "total fl"})
        if group == None:
            print('不存在页面')
        elif blacklist.nameList(unit):
            print('不存在房源')
        # 出现拼写错误时，取第一个跳转链接
        elif len(soup.find_all('div', {'class','spellcheck'})) != 0:
            print('跳转')
            # spellGuess = soup.find_all('div', {'class','spellcheck'})[0]
            # ele = spellGuess.find_all('a')[0]
            # getApartmentInfo(ele.attrs['data-el'])
        # 找到0套待售房源    
        elif group.find_all('span')[0].text.strip(' ') == '0':
            print('无')
        else:
            g_data = soup.find_all('div', {'class': 'info clear'})
            detail = formatInfo(g_data, unit, param['name'])
            write_excel(detail)



# 格式化返回房源信息
def formatInfo(g_data, community, schoolName):
    list = []
    for item in g_data:
        if community in item.find_all('a')[1].text:
            list.append(getUploadInfo(item.find('a').attrs['href'], schoolName))
    return list        

            
# 写入excel
def write_excel(info):
    print(info)
    ws.append([
        'schoolName',  # 学校名称
        'total',       # 总价
        'unit',        # 单价
        'room',        # 户型
        'direction',   # 朝向
        'community',   # 小区
        'age',         # 房龄
        'size',        # 平方米
        'url',         # 链接
    ])

    for key in range(len(info)):
        ws.cell(row = (key + 2), column = 1).value = info[key]['schoolName']
        ws.cell(row = (key + 2), column = 2).value = info[key]['total']
        ws.cell(row = (key + 2), column = 3).value = info[key]['unit']
        ws.cell(row = (key + 2), column = 4).value = info[key]['room']
        ws.cell(row = (key + 2), column = 5).value = info[key]['direction']
        ws.cell(row = (key + 2), column = 6).value = info[key]['community']
        ws.cell(row = (key + 2), column = 7).value = info[key]['age']
        ws.cell(row = (key + 2), column = 8).value = info[key]['size']
        ws.cell(row = (key + 2), column = 9).value = info[key]['url']
    wb.save('futian' + '.xlsx')


getTopElementrySchool('futian', [400, 550])



# def test():
#     url = 'https://sz.lianjia.com/ershoufang/bp500ep550rs'+ '花好园/'
#     r = requests.get(url)
#     soup = BeautifulSoup(r.content, 'lxml')
    
#     group = soup.find("h2", {"class", "total fl"})
    
#     # 出现拼写错误时，取第一个跳转链接
#     if len(soup.find_all('div', {'class','spellcheck'})) != 0:
#         spellGuess = soup.find_all('div', {'class','spellcheck'})[0]
#         ele = spellGuess.find_all('a')[0]
#         print(ele.attrs['data-el'])
#         getApartmentInfo(ele.attrs['data-el'], [400, 550])
#     # 找到0套待售房源    
#     elif group.find_all('span')[0].text.strip(' ') == '0':
#         print('无')
#     else:
#         g_data = soup.find_all('div', {'class': 'info clear'})
#         formatInfo(g_data, '花好园')




# test()



